#!/usr/bin/env python3
"""Sync CAI sprint tickets to local Excel/CSV and Google Sheets.

Usage:
  python scripts/sync_cai_sprint_sheet.py              # xlsx + csv only
  python scripts/sync_cai_sprint_sheet.py --google     # also push to Google Sheets

Google Sheets setup (one-time):
  1. Create a GCP service account with Sheets API enabled.
  2. Download JSON key → save as secrets/google-sheets-sa.json (gitignored).
  3. Share the spreadsheet with the service account email (Editor).
  4. Run: GOOGLE_APPLICATION_CREDENTIALS=secrets/google-sheets-sa.json \\
         python scripts/sync_cai_sprint_sheet.py --google
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import sys
from datetime import datetime, timezone
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO_ROOT / "scripts"))

from cai_sprint_data import TRACKER_JSON, parse_tickets, summary_counts  # noqa: E402

HEADERS = [
    "ID",
    "Week",
    "Section",
    "Title",
    "Size",
    "Depends",
    "Description / Acceptance",
    "Completion",
    "Blockers",
]

XLSX_PATH = REPO_ROOT / "docs" / "CAI_SPRINT_PLAN.xlsx"
CSV_PATH = REPO_ROOT / "docs" / "CAI_SPRINT_PLAN.csv"


def ticket_rows(tickets) -> list[list[str]]:
    return [
        [
            t.id,
            t.week,
            t.section,
            t.title,
            t.size,
            t.depends,
            t.description,
            t.completion,
            t.blockers,
        ]
        for t in tickets
    ]


def write_csv(tickets) -> None:
    with CSV_PATH.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(HEADERS)
        writer.writerows(ticket_rows(tickets))
    print(f"Wrote {CSV_PATH}")


def write_xlsx(tickets) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "CAI Tickets"
    ws.append(HEADERS)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    fills = {
        "Done": PatternFill("solid", fgColor="C6EFCE"),
        "Done (partial)": PatternFill("solid", fgColor="FFEB9C"),
        "Deferred": PatternFill("solid", fgColor="E7E6E6"),
    }

    for row in ticket_rows(tickets):
        ws.append(row)
        r = ws.max_row
        for col in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=r, column=col)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if col == 8:
                cell.fill = fills.get(row[7], PatternFill())

    counts = summary_counts(tickets)
    ws2 = wb.create_sheet("Summary")
    ws2.append(["Metric", "Value"])
    for key, value in [
        ("Total tickets", len(tickets)),
        ("Done", counts.get("Done", 0)),
        ("Done (partial)", counts.get("Done (partial)", 0)),
        ("Not started", counts.get("Not started", 0)),
        ("Deferred", counts.get("Deferred", 0)),
        ("Generated (UTC)", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M")),
        ("Source", "docs/CAI_SPRINT_PLAN.md + docs/cai_sprint_tracker.json"),
    ]:
        ws2.append([key, value])

    widths = [10, 10, 28, 42, 6, 18, 60, 14, 40]
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{ws.max_row}"

    wb.save(XLSX_PATH)
    print(f"Wrote {XLSX_PATH}")


def push_google_sheets(tickets) -> None:
    creds_path = os.environ.get("GOOGLE_APPLICATION_CREDENTIALS")
    if not creds_path or not Path(creds_path).exists():
        raise SystemExit(
            "Set GOOGLE_APPLICATION_CREDENTIALS to a service-account JSON file.\n"
            "See scripts/sync_cai_sprint_sheet.py docstring."
        )

    import gspread
    from google.oauth2.service_account import Credentials

    tracker = json.loads(TRACKER_JSON.read_text(encoding="utf-8"))
    spreadsheet_id = tracker["spreadsheet_id"]
    worksheet_gid = int(tracker["worksheet_gid"])

    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]
    credentials = Credentials.from_service_account_file(creds_path, scopes=scopes)
    client = gspread.authorize(credentials)
    spreadsheet = client.open_by_key(spreadsheet_id)
    worksheet = spreadsheet.get_worksheet_by_id(worksheet_gid)

    values = [HEADERS] + ticket_rows(tickets)
    worksheet.clear()
    worksheet.update(values, value_input_option="USER_ENTERED")

    # Summary tab (create or replace)
    summary_name = "Summary"
    try:
        summary_ws = spreadsheet.worksheet(summary_name)
        summary_ws.clear()
    except gspread.WorksheetNotFound:
        summary_ws = spreadsheet.add_worksheet(summary_name, rows=20, cols=2)

    counts = summary_counts(tickets)
    summary_values = [
        ["Metric", "Value"],
        ["Total tickets", len(tickets)],
        ["Done", counts.get("Done", 0)],
        ["Done (partial)", counts.get("Done (partial)", 0)],
        ["Not started", counts.get("Not started", 0)],
        ["Deferred", counts.get("Deferred", 0)],
        ["Last sync (UTC)", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M")],
    ]
    summary_ws.update(summary_values, value_input_option="USER_ENTERED")

    tracker["last_synced"] = datetime.now(timezone.utc).isoformat()
    TRACKER_JSON.write_text(json.dumps(tracker, indent=2) + "\n", encoding="utf-8")
    print(f"Pushed {len(tickets)} tickets to Google Sheet {spreadsheet_id} (gid={worksheet_gid})")


def main() -> None:
    parser = argparse.ArgumentParser(description="Sync CAI sprint tracker exports")
    parser.add_argument("--google", action="store_true", help="Push to Google Sheets")
    args = parser.parse_args()

    tickets = parse_tickets()
    if len(tickets) != 83:
        raise SystemExit(f"Expected 83 tickets, parsed {len(tickets)}")

    write_csv(tickets)
    write_xlsx(tickets)

    if args.google:
        push_google_sheets(tickets)
    else:
        print("Skipped Google Sheets (pass --google after setting credentials).")


if __name__ == "__main__":
    main()
